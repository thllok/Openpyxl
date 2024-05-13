import os

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import *
from openpyxl import load_workbook, Workbook
def formatting():
    global font_bold,font_notbold,border,purpleFill,yellowFill,orangeFill,no_fill,no_border,GrayFill,GreenFill,rule_databar
    no_fill = PatternFill(fill_type=None)
    side = Side(border_style=None)
    no_border = borders.Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )
    font_bold=Font(name="Arial",size=11,bold=True,color="000000")
    font_notbold=Font(name="Arial",size=11,bold=False,color="000000")
    thin=Side(border_style="thin", color="000000")
    double=Side(border_style="double", color="000000")
    border=Border(top=thin, left=thin, right=thin, bottom=thin)

    yellow="FFFF99"  ##YELLOW
    yellowFill=PatternFill(start_color=yellow,
                             end_color=yellow,
                             fill_type='solid')

    purple='9999FF'
    purpleFill=PatternFill(start_color=purple,
                             end_color=purple,
                             fill_type='solid')

    orange="FFC125"  ##orange
    OrangeFill=PatternFill(start_color=orange,
                             end_color=orange,
                             fill_type='solid')
    gray="E7E6E6"  ##gray
    GrayFill=PatternFill(start_color=gray,
                             end_color=gray,
                             fill_type='solid')
    green="B8C2AD"  ##green

    GreenFill=PatternFill(start_color=green,
                             end_color=green,
                             fill_type='solid')

def reduce_excel_col_name(n):
    result = ''

    while n > 0:
        index = (n - 1) % 26
        result += chr(index + ord('A'))
        n = (n - 1) // 26

    return result[::-1]
if __name__ =="__main__":
    formatting()
    file_name = os.getcwd()+"\\testing.xlsx"
    wb=load_workbook(file_name,data_only = False)
    # or create a new excel file
    wb=Workbook()
    ws=wb[wb.sheetnames[0]]
    #change the tab name
    ws.title = "testing"
    ws.cells(row= 1,column=1).value ="testing"
    #or
    ws[reduce_excel_col_name(1)+str(1)].value = "testing"
    ws[reduce_excel_col_name(1)+str(1)].font = font_bold
    ws[reduce_excel_col_name(1) + str(1)].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    #horizontal = "center" or "left" or "right"
    #vertical = "center" or "bottom" or "top"
    ws[reduce_excel_col_name(1) + str(1)].fill = yellowFill

    # way to formatting range of cells
    for row in ws.iter_rows('A1:D100'):
        for cell in row:
            cell.value = "testing"
            cell.font = font_bold
            cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
            cell.fill = yellowFill
    # paste a dataframe in the excel
    # if the dataframe is called "benchmark"
    benchmark = pd.DataFrame()
    rows = dataframe_to_rows(benchmark, index=False, header=True)
    for r_idx, row in enumerate(rows, 3):
        for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(file_name)